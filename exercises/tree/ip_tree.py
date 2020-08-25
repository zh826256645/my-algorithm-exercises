# -*- encoding: utf-8 -*-
"""
IP 查找树的练习

参照 https://github.com/wzhe06/ipdatabase

@File    :   ip_tree.py
@Time    :   2020/08/24 17:43:45
@Author  :   Zhong Hao
@Version :   1.0
@Contact :   zh826256645@gmail.com
@Desc    :   None
"""
import csv
import time

from openpyxl import load_workbook

from model.ip_tree import IPTree


IP_CSV_PATH = '/Users/zhonghao/Projects/my-algorithm-exercises/static/files/ipDatabase.csv'
REGION_EXCEL_PATH = '/Users/zhonghao/Projects/my-algorithm-exercises/static/files/ipRegion.xlsx'


class IPHelper:

    ip_tree = None
    region_map = None

    @staticmethod
    def build_tree(ip_csv_path):
        """构建 IP 树

        Keyword arguments:
        ip_csv_path -- csv 文件路径
        """
        print('开始构建 IP 树')
        t1 = time.time()
        IPHelper.ip_tree = IPTree()
        with open(ip_csv_path, newline='') as csvfile:
            reader = csv.reader(csvfile)
            i = 1
            for row in reader:
                ip_start, ip_end, address_code = row[0], row[1], row[2]
                IPHelper.ip_tree.train(ip_start, ip_end, address_code)
                i += 1

        print(f'构建 IP 树所用时间：{time.time() - t1}')

    @staticmethod
    def load_region(region_excel_path):
        """导入城市信息

        Keyword arguments:
        argument -- description
        Return:
        """
        print('开始载入城市信息')
        t1 = time.time()

        IPHelper.region_map = dict()
        region_sheet = load_workbook(region_excel_path)['地域信息表']
        for row in region_sheet.iter_rows():
            upper_region, region, address_code = row[0].value, row[1].value, str(row[2].value)
            if address_code.isdigit():
                IPHelper.region_map[str(address_code)] = {'upper_region': upper_region, 'region': region, 'address_code': address_code}

        print(f'导入城市信息所用时间：{time.time() - t1}')

    @staticmethod
    def start(ip_csv_path: str, region_excel_path: str):
        """开始程序

        Keyword arguments:
        ip_csv_path -- IP 数据文件
        Return: return_description
        """
        IPHelper.build_tree(ip_csv_path)
        IPHelper.load_region(region_excel_path)

        while True:
            ip = input('请输入 IP 地址: ')
            info = IPHelper.find(ip)
            if info:
                print('您查找的 IP 信息如下')
                print(f'上级地域: {info["upper_region"]}, 地域: {info["region"]}')

    @staticmethod
    def find(ip: str) -> dict():
        """"查找 IP 对应的信息

        Keyword arguments:
        ip -- IP 地址
        Return: IP 对应的信息
        """
        print(f'开始查询 IP({ip}) 信息')
        t1 = time.time()
        address_code = IPHelper.ip_tree.find_ip(ip)
        print(f'查询成功，所用时间: {time.time() - t1}')

        if address_code and address_code.isdigit():
            return IPHelper.region_map[address_code]
        return None


def main():
    IPHelper.start(IP_CSV_PATH, REGION_EXCEL_PATH)


if __name__ == "__main__":
    main()
